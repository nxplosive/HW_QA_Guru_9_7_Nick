import os
import csv
import zipfile
import openpyxl
import test_download_files
from zipfile import ZipFile
from pypdf import PdfReader
from io import TextIOWrapper
from script_os import ARCHIVE, TMP_DIR, FILES_LIST


# Скачивание текстовых файлов в папку tmp
def test_tmp_download():
    test_download_files.test_download_pdf()
    test_download_files.test_download_xlsx()
    test_download_files.test_download_csv()


# Создание папки resources и архива в ней
def test_zip_creation(zip_creation):
    with ZipFile(ARCHIVE, "w", compression=zipfile.ZIP_DEFLATED) as zfile:
        for file in FILES_LIST:
            add_file = os.path.join(TMP_DIR, file)
            zfile.write(add_file, os.path.basename(add_file))


# Проверка содержимого архива
def test_assertion_zip():
    with zipfile.ZipFile('resources/archive.zip', "r") as zf:
        file_list_zip = zf.namelist()
    assert file_list_zip == FILES_LIST


# Проверка содержимого PDF
def test_assertion_pdf():
    with ZipFile('resources/archive.zip') as zf:
        with zf.open('pdf_file.pdf') as pdf:
            pdf_reader = PdfReader(pdf)
            assert "In our models, we have a standard" in pdf_reader.pages[10].extract_text()


# Проверка содержимого CSV
def test_assertion_csv():
    with ZipFile('resources/archive.zip') as zf:
        with zf.open('csv_file.csv', "r") as csv_file:
            csv_reader = list(csv.reader(TextIOWrapper(csv_file, 'utf-8')))
            assert "Game Number" == csv_reader[0][0]
            assert " 16" == csv_reader[4][1]


# Проверка содержимого XLSX
def test_assertion_xlsx():
    with ZipFile('resources/archive.zip') as zf:
        with zf.open('xlsx_file.xlsx', "r") as xlsx:
            wb = openpyxl.load_workbook(xlsx)
            sheet = wb.active
            first = sheet.cell(row=1, column=1).value
            second = sheet.cell(row=2, column=2).value
            assert first == "ID", f"1 is not xlsx"
            assert second == "January", f"1 is not xlsx"
